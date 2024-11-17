import time
import psutil
import clickhouse_connect
import psycopg2
from openpyxl import Workbook, load_workbook

# Konfiguracja połączenia z ClickHouse
CLICKHOUSE_HOST = 'localhost'
CLICKHOUSE_PORT = 8123
CLICKHOUSE_USER = 'default'
CLICKHOUSE_PASSWORD = '1234'
CLICKHOUSE_DATABASE = 'default'

# Konfiguracja połączenia z PostgreSQL
PG_HOST = 'localhost'
PG_PORT = 5432
PG_USER = 'postgres'
PG_PASSWORD = '1234'
PG_DATABASE = 'AirFlights'

# Funkcja monitorująca zużycie CPU i RAM przed, w trakcie i po zapytaniu
def monitor_resources():
    cpu_usage = psutil.cpu_percent(interval=1)
    memory_usage = psutil.virtual_memory().percent
    return cpu_usage, memory_usage

# Funkcja wykonująca zapytanie w ClickHouse i mierząca czas oraz zasoby
def execute_clickhouse_query(query):
    client = clickhouse_connect.get_client(
        host=CLICKHOUSE_HOST, port=CLICKHOUSE_PORT,
        username=CLICKHOUSE_USER, password=CLICKHOUSE_PASSWORD
    )

    # Pomiar czasu połączenia
    connect_start = time.time()
    client.ping()
    connect_time = time.time() - connect_start

    # Pomiar czasu zapytania
    cpu_before, memory_before = monitor_resources()
    query_start = time.time()
    result = client.query(query).result_rows
    query_time = time.time() - query_start
    cpu_after, memory_after = monitor_resources()

    client.close()

    return {
        "result": result,
        "execution_time": query_time,  # Czas samego zapytania
        "connect_time": connect_time,  # Czas na połączenie
        "cpu_before": cpu_before,
        "cpu_after": cpu_after,
        "memory_before": memory_before,
        "memory_after": memory_after
    }


# Funkcja wykonująca zapytanie w PostgreSQL i mierząca czas oraz zasoby
def execute_pg_query(query):
    # Opóźnienie o 5 sekund przed wykonaniem zapytania
    time.sleep(10)

    # Mierzymy czas połączenia
    connect_start = time.time()
    conn = psycopg2.connect(
        host=PG_HOST, port=PG_PORT,
        user=PG_USER, password=PG_PASSWORD,
        database=PG_DATABASE
    )
    connect_time = time.time() - connect_start

    cursor = conn.cursor()

    # Mierzymy czas wysyłania zapytania
    cpu_before, memory_before = monitor_resources()
    query_start = time.time()

    cursor.execute(query)
    result = cursor.fetchall()
    query_time = time.time() - query_start

    cpu_after, memory_after = monitor_resources()
    cursor.close()
    conn.close()

    return {
        "result": result,
        "execution_time": query_time,  # Czas samego zapytania
        "connect_time": connect_time,  # Czas na połączenie
        "cpu_before": cpu_before,
        "cpu_after": cpu_after,
        "memory_before": memory_before,
        "memory_after": memory_after
    }


# Funkcja zapisująca wyniki do pliku Excel
# Funkcja zapisująca wyniki do pliku Excel
def load_queries_from_file(file_path):
    """
    Wczytuje zapytania z pliku i dzieli je na grupy na podstawie nagłówków.
    """
    queries = {}
    current_group = None

    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith("[") and line.endswith("]"):
                current_group = line[1:-1]
                queries[current_group] = []
            elif line and current_group:
                queries[current_group].append(line)

    return queries


def execute_queries_by_group(queries, db_type="ClickHouse"):
    """
    Wykonuje zapytania podzielone na grupy dla podanej bazy danych (ClickHouse lub PostgreSQL).
    """
    results = []

    for group, query_list in queries.items():
        print(f"\nExecuting queries in group: {group}")

        for query in query_list:
            print(f"Executing: {query}")
            if db_type == "ClickHouse":
                result = execute_clickhouse_query(query)
            elif db_type == "PostgreSQL":
                result = execute_pg_query(query)
            else:
                raise ValueError("Unsupported database type.")

            # Dodaj wyniki do listy
            results.append((group, db_type, result))
            save_results_to_excel(result, group, db_type)

    return results


def save_results_to_excel(results, group, db_type):
    """
    Zapisuje wyniki do pliku Excel z podziałem na grupy.
    """
    try:
        wb = load_workbook("query_comparison.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Comparison"
        ws.append([
            "Group", "Database", "Execution Time (s)", "Connection Time (s)",
            "CPU Before (%)", "CPU After (%)",
            "Memory Before (%)", "Memory After (%)", "Result Rows"
        ])

    # Dodaj wyniki do arkusza
    ws.append([
        group, db_type, results["execution_time"], results["connect_time"],
        results["cpu_before"], results["cpu_after"],
        results["memory_before"], results["memory_after"],
        len(results["result"])
    ])

    wb.save("query_comparison.xlsx")
    print(f"Wyniki zapisane w pliku query_comparison.xlsx dla grupy {group} ({db_type}).")


# Wczytanie zapytań z pliku
queries = load_queries_from_file("queries.txt")

# Wykonanie zapytań w ClickHouse
execute_queries_by_group(queries, db_type="ClickHouse")

# Wykonanie zapytań w PostgreSQL
# execute_queries_by_group(queries, db_type="PostgreSQL")

